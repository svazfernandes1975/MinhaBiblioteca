#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
COLETOR DE CAPAS v5 — BUSCA GERAL NA WEB

A v5 mantém as capas já encontradas pela v4 e trabalha principalmente
nos livros que ficaram sem capa.

Além de Google Books/Open Library, usa resultados de busca na web para
encontrar páginas de livros em:
- Skoob
- Estante Virtual
- sebos/livrarias
- Mercado Livre
- Amazon
- Goodreads
- páginas de editoras/livrarias
- outras páginas indexadas

A página encontrada é aberta e o programa procura a imagem principal
da capa em:
- og:image
- twitter:image
- JSON-LD image
- imagens HTML

A busca é feita primeiro pelo ISBN exato; depois por título+autor e por
variações simplificadas do título.

IMPORTANTE:
Este programa é deliberadamente conservador: não salva uma imagem só
porque ela apareceu na busca. Ele exige sinais de que a página corresponde
ao livro (ISBN, título ou título+autor) e registra a página de origem.
"""

import html
import json
import re
import time
import zipfile
from pathlib import Path
from urllib.parse import quote, urljoin, urlparse, unquote

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

INPUT = Path("Lista_de_Livros_com_genero.xlsx")
V4_REPORT = Path("resultado_capas_v4.xlsx")
V4_DIR = Path("capas_livros_v4")

OUTPUT_DIR = Path("capas_livros_v5")
ZIP_FILE = Path("capas_livros_v5.zip")
REPORT_FILE = Path("resultado_capas_v5.xlsx")

GB_URL = "https://www.googleapis.com/books/v1/volumes"
OL_SEARCH = "https://openlibrary.org/search.json"
OL_COVER_ID = "https://covers.openlibrary.org/b/id/{cover_id}-L.jpg"

session = requests.Session()
session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
})

# Domínios que costumam conter páginas de livros/capas.
PREFERRED_DOMAINS = [
    "skoob.com.br",
    "estantevirtual.com.br",
    "sebodomessias.com.br",
    "portaldoslivreiros.com.br",
    "mercadolivre.com.br",
    "amazon.com.br",
    "goodreads.com",
    "livrariacultura.com.br",
    "travessa.com.br",
    "livrariadavila.com.br",
    "companhiadasletras.com.br",
    "record.com.br",
    "intrinseca.com.br",
    "rocco.com.br",
    "harpercollins.com.br",
    "editoraarqueiro.com.br",
    "globolivros.com.br",
    "editoraplaneta.com.br",
    "martinclaret.com.br",
    "bertrandbrasil.com.br",
]

def clean_isbn(v):
    if v is None:
        return ""
    return re.sub(r"[^0-9Xx]", "", str(v).strip()).upper()

def isbn_variants(isbn):
    isbn = clean_isbn(isbn)
    out = []
    if isbn:
        out.append(isbn)
    if len(isbn) == 13 and isbn.startswith("978") and isbn[:-1].isdigit():
        core = isbn[3:-1]
        s = sum(int(ch) * (1 if i % 2 == 0 else 3)
                for i, ch in enumerate(core))
        out.append(core + str((10 - s % 10) % 10))
    elif len(isbn) == 10 and isbn[:9].isdigit():
        core = "978" + isbn[:9]
        s = sum(int(ch) * (1 if i % 2 == 0 else 3)
                for i, ch in enumerate(core))
        out.append(core + str((10 - s % 10) % 10))
    return list(dict.fromkeys(out))

def norm(s):
    s = str(s or "").lower()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()

def title_variants(title):
    original = str(title or "").strip()
    out = []

    def add(x):
        x = re.sub(r"\s+", " ", str(x or "")).strip(" -–—:;,.")
        if len(x) >= 2 and x not in out:
            out.append(x)

    add(original)

    # Remove parenthetical/bracket series information.
    s = re.sub(
        r"\s*[\(\[][^)\]]*(?:#\s*\d+|n[ºo°]?\s*\d+|volume|vol\.?|book|livro|"
        r"s[ée]rie|series|trilogy|trilogia|saga)[^)\]]*[\)\]]",
        "",
        original,
        flags=re.I,
    )
    add(s)

    # Anything before a parenthesis/bracket is usually the clean title.
    m = re.search(r"\s*[\(\[]", original)
    if m:
        add(original[:m.start()])

    # "Título - (Série, #5)" and similar.
    for sep in (" - (", " – (", " — (", " - [", " – [", " — ["):
        if sep in original:
            add(original.split(sep, 1)[0])

    # Remove final #number.
    add(re.sub(r"\s*[\(\[]?\s*#\s*\d+\s*[\)\]]?\s*$", "", original))

    # Remove explicit volume/book number at end.
    add(re.sub(
        r"\s*[-–—:;,]?\s*[\(\[]?\s*"
        r"(?:livro|book|volume|vol\.?|n[ºo°]?|#)\s*\d+"
        r"[^\)\]]*[\)\]]?\s*$",
        "",
        original,
        flags=re.I,
    ))

    return out

def title_score(wanted, got):
    a, b = norm(wanted), norm(got)
    if not a or not b:
        return 0
    if a == b:
        return 1.0
    if a in b or b in a:
        return .95
    wa, wb = set(a.split()), set(b.split())
    return len(wa & wb) / max(1, len(wa | wb))

def author_score(wanted, got):
    if not wanted or not got:
        return 0

    def lastnames(s):
        vals = []
        for x in re.split(r"[,;&/]| and | e ", str(s), flags=re.I):
            p = norm(x).split()
            if p:
                vals.append(p[-1])
        return vals

    a = lastnames(wanted)
    b = []
    for x in got:
        p = norm(x).split()
        if p:
            b.append(p[-1])

    if not a or not b:
        return 0
    return max(1 if x == y else 0 for x in a for y in b)

def google_search(query):
    try:
        r = session.get(
            GB_URL,
            params={"q": query, "maxResults": 40},
            timeout=25,
        )
        if r.status_code != 200:
            return []
        return r.json().get("items", [])
    except Exception:
        return []

def google_cover(title, author, isbn):
    variants = title_variants(title)
    isbnset = set(isbn_variants(isbn))

    # ISBN
    for v in isbn_variants(isbn):
        for item in google_search(f"isbn:{v}"):
            vi = item.get("volumeInfo", {})
            links = vi.get("imageLinks") or {}
            if not links:
                continue
            ids = {
                clean_isbn(x.get("identifier"))
                for x in vi.get("industryIdentifiers", [])
            }
            if isbnset & ids:
                return best_google_image(vi), "Google Books / ISBN"

    # Title variants
    best = None
    best_score = -1
    for t in variants:
        for q in (
            f'intitle:"{t}" inauthor:"{author}"',
            f'"{t}" "{author}"',
            f'intitle:"{t}"',
        ):
            for item in google_search(q):
                vi = item.get("volumeInfo", {})
                links = vi.get("imageLinks") or {}
                if not links:
                    continue
                ts = max(title_score(x, vi.get("title","")) for x in variants)
                aus = author_score(author, vi.get("authors", []))
                score = 65 * ts + 45 * aus
                if score > best_score:
                    best = item
                    best_score = score

    if best and best_score >= 50:
        return best_google_image(best["volumeInfo"]), "Google Books / título"

    return None, ""

def best_google_image(vi):
    links = vi.get("imageLinks") or {}
    for key in ("extraLarge", "large", "medium", "small", "thumbnail"):
        if links.get(key):
            return links[key].replace("http://", "https://")
    return None

def ol_search(params):
    try:
        r = session.get(OL_SEARCH, params=params, timeout=25)
        if r.status_code != 200:
            return []
        return r.json().get("docs", [])
    except Exception:
        return []

def openlibrary_cover(title, author, isbn):
    variants = title_variants(title)
    isbnset = set(isbn_variants(isbn))

    # ISBN
    for v in isbn_variants(isbn):
        for d in ol_search({"isbn": v, "limit": 40}):
            if not d.get("cover_i"):
                continue
            ids = set()
            for k in ("isbn", "isbn10", "isbn13"):
                vals = d.get(k) or []
                if isinstance(vals, str):
                    vals = [vals]
                ids.update(clean_isbn(x) for x in vals)
            if isbnset & ids:
                return OL_COVER_ID.format(cover_id=d["cover_i"]), "Open Library / ISBN"

    # title/author
    best = None
    best_score = -1
    for t in variants:
        for params in (
            {"title": t, "author": author, "limit": 40},
            {"q": f'"{t}" "{author}"', "limit": 40},
            {"title": t, "limit": 40},
        ):
            for d in ol_search(params):
                if not d.get("cover_i"):
                    continue
                ts = max(title_score(x, d.get("title","")) for x in variants)
                aus = author_score(author, d.get("author_name", []))
                score = 65 * ts + 45 * aus
                if score > best_score:
                    best = d
                    best_score = score

    if best and best_score >= 50:
        return OL_COVER_ID.format(cover_id=best["cover_i"]), "Open Library / título"

    return None, ""

def bing_search(query):
    """Return result URLs from Bing HTML search."""
    url = "https://www.bing.com/search"
    try:
        r = session.get(
            url,
            params={"q": query, "count": 20, "setlang": "pt-BR"},
            timeout=25,
        )
        if r.status_code != 200:
            return []
        soup = BeautifulSoup(r.text, "html.parser")
        urls = []
        for li in soup.select("li.b_algo"):
            a = li.select_one("h2 a")
            if a and a.get("href"):
                urls.append(a["href"])
        # Fallback: collect normal links.
        if not urls:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("http") and "bing.com" not in href:
                    urls.append(href)
        return list(dict.fromkeys(urls))
    except Exception:
        return []

def google_web_search(query):
    """Fallback search parser for Google HTML results."""
    try:
        r = session.get(
            "https://www.google.com/search",
            params={"q": query, "num": 20, "hl": "pt-BR"},
            timeout=25,
        )
        if r.status_code != 200:
            return []
        soup = BeautifulSoup(r.text, "html.parser")
        urls = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("/url?q="):
                href = href.split("/url?q=", 1)[1].split("&", 1)[0]
            if href.startswith("http") and "google.com" not in href:
                urls.append(href)
        return list(dict.fromkeys(urls))
    except Exception:
        return []

def domain(url):
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""

def preferred_domain_score(url):
    d = domain(url)
    for i, x in enumerate(PREFERRED_DOMAINS):
        if d == x or d.endswith("." + x):
            return 30 - min(i, 20)
    return 0

def page_matches(text, title, author, isbn):
    t = norm(text)
    variants = title_variants(title)

    isbn_ok = any(v and v in re.sub(r"[^0-9Xx]", "", text).upper()
                  for v in isbn_variants(isbn))

    ts = max((title_score(v, text[:10000]) for v in variants), default=0)
    # More reliable lexical title presence:
    title_present = any(norm(v) in t for v in variants if len(norm(v)) >= 4)
    author_present = norm(author) in t if author else False

    score = preferred_domain_score(text if False else "")
    score = 0

    if isbn_ok:
        score += 100
    if title_present:
        score += 55
    if author_present:
        score += 30

    # Loose token overlap for pages that use punctuation/HTML entities.
    for v in variants:
        w = set(norm(v).split())
        if len(w) >= 2:
            overlap = sum(1 for x in w if x in t) / len(w)
            score += 25 * overlap
            break

    return score

def extract_image_urls(page_url, html_text):
    soup = BeautifulSoup(html_text, "html.parser")
    candidates = []

    def add(u, source):
        if not u:
            return
        u = html.unescape(str(u)).strip()
        if u.startswith("//"):
            u = "https:" + u
        u = urljoin(page_url, u)
        if u.startswith("http"):
            candidates.append((u, source))

    # Open Graph / Twitter.
    for meta in soup.find_all("meta"):
        prop = (meta.get("property") or meta.get("name") or "").lower()
        content = meta.get("content")
        if prop in ("og:image", "og:image:url", "twitter:image", "twitter:image:src"):
            add(content, prop)

    # JSON-LD.
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(tag.string or tag.get_text())
        except Exception:
            continue

        def walk(x):
            if isinstance(x, dict):
                img = x.get("image")
                if isinstance(img, str):
                    add(img, "jsonld:image")
                elif isinstance(img, dict):
                    add(img.get("url"), "jsonld:image")
                elif isinstance(img, list):
                    for y in img:
                        if isinstance(y, str):
                            add(y, "jsonld:image")
                        elif isinstance(y, dict):
                            add(y.get("url"), "jsonld:image")
                for v in x.values():
                    if isinstance(v, (dict, list)):
                        walk(v)
            elif isinstance(x, list):
                for v in x:
                    walk(v)

        walk(data)

    # HTML images. Prefer large-looking images and class/alt terms around cover.
    for img in soup.find_all("img"):
        alt = norm(img.get("alt", ""))
        cls = norm(" ".join(img.get("class", [])))
        src = img.get("src") or img.get("data-src") or img.get("data-original")
        if src:
            if any(x in (alt + " " + cls) for x in (
                "capa", "cover", "livro", "book", "produto"
            )):
                add(src, "img:cover")
            else:
                add(src, "img")

    # De-duplicate while keeping order.
    out = []
    seen = set()
    for u, source in candidates:
        if u not in seen:
            seen.add(u)
            out.append((u, source))
    return out

def image_quality_score(url, source):
    u = url.lower()
    score = 0

    if source.startswith("og:") or source.startswith("twitter:"):
        score += 35
    elif source.startswith("jsonld"):
        score += 30
    elif source == "img:cover":
        score += 25

    # Prefer likely full-size image URLs.
    for word, pts in (
        ("large", 15),
        ("original", 20),
        ("full", 15),
        ("master", 15),
        ("1200", 12),
        ("1000", 10),
        ("800", 8),
        ("600", 5),
    ):
        if word in u:
            score += pts

    # Avoid obvious thumbnails.
    for word in ("thumb", "thumbnail", "tiny", "small", "avatar", "logo"):
        if word in u:
            score -= 15

    return score

def download_image(url, path):
    try:
        r = session.get(url, timeout=35)
        if r.status_code != 200:
            return False, f"HTTP {r.status_code}"
        data = r.content
        ct = r.headers.get("content-type", "").lower()
        if len(data) < 5000:
            return False, "imagem muito pequena"
        if "image" not in ct and not data.startswith(
            (b"\xff\xd8", b"\x89PNG", b"RIFF")
        ):
            return False, "não parece imagem"
        path.write_bytes(data)
        return True, ""
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"

def web_cover(title, author, isbn):
    variants = title_variants(title)

    queries = []

    # Exact ISBN gets top priority.
    for v in isbn_variants(isbn):
        queries.extend([
            f'"{v}" capa livro',
            f'"{v}" "{variants[0]}"',
        ])

    # Then clean title + author.
    for t in variants:
        queries.extend([
            f'"{t}" "{author}" capa livro',
            f'"{t}" "{author}" livro',
            f'"{t}" capa',
        ])

    seen_urls = set()
    pages = []

    for q in queries:
        urls = bing_search(q)
        if not urls:
            urls = google_web_search(q)

        for u in urls:
            if u in seen_urls:
                continue
            seen_urls.add(u)
            d = domain(u)

            # Prefer book-related domains but don't exclude unknown domains.
            pages.append((preferred_domain_score(u), u))

        # Search engines can be rate-limited; don't hammer them.
        time.sleep(.6)

    pages.sort(reverse=True)

    best = None
    best_score = -1

    # Inspect a reasonable number of pages.
    for _, url in pages[:45]:
        try:
            r = session.get(url, timeout=25)
            if r.status_code != 200:
                continue
            ct = r.headers.get("content-type", "").lower()
            if "html" not in ct:
                continue

            text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)

            match = page_matches(text, title, author, isbn)
            match += preferred_domain_score(url)

            if match < 35:
                continue

            imgs = extract_image_urls(url, r.text)
            if not imgs:
                continue

            for img_url, src in imgs:
                q = match + image_quality_score(img_url, src)

                # A page with exact ISBN is much safer than title-only.
                if q > best_score:
                    best = (img_url, url, q, src)
                    best_score = q

        except Exception:
            continue

        time.sleep(.35)

    if best:
        return best

    return None

def find_columns(ws):
    headers = [str(c.value or "").strip() for c in ws[1]]

    def find(names):
        for name in names:
            for i, h in enumerate(headers):
                if norm(h) == norm(name):
                    return i
        return None

    ti = find(["Título", "Titulo", "Title"])
    ai = find(["Autor", "Author"])
    ii = find(["ISBN", "ISBN13", "ISBN-13", "ISBN10", "ISBN-10"])

    if ti is None or ai is None:
        raise RuntimeError(f"Colunas encontradas: {headers}")

    return ti, ai, ii

def load_v4_results():
    if not V4_REPORT.exists():
        return {}

    try:
        wb = load_workbook(V4_REPORT, data_only=True)
        ws = wb["Resultado"]
        rows = list(ws.values)
        if not rows:
            return {}
        headers = rows[0]
        out = {}
        for row in rows[1:]:
            d = dict(zip(headers, row))
            out[d.get("Nº")] = d
        return out
    except Exception:
        return {}

def copy_existing_covers(v4):
    OUTPUT_DIR.mkdir(exist_ok=True)

    if not V4_DIR.exists():
        return

    for d in v4.values():
        if d.get("Status") != "Encontrada":
            continue
        name = d.get("Arquivo")
        if not name:
            continue
        src = V4_DIR / name
        dst = OUTPUT_DIR / name
        if src.exists() and not dst.exists():
            try:
                dst.write_bytes(src.read_bytes())
            except Exception:
                pass

def main():
    print("=" * 72)
    print("COLETOR DE CAPAS v5 — BUSCA GERAL NA WEB")
    print("=" * 72)
    print()

    if not INPUT.exists():
        print(f"ERRO: não encontrei {INPUT.resolve()}")
        input("Enter para sair...")
        return

    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = load_workbook(INPUT, data_only=True)
    ws = wb.active
    ti, ai, ii = find_columns(ws)

    v4 = load_v4_results()
    copy_existing_covers(v4)

    total = ws.max_row - 1
    results = []

    # First pass: preserve the v4 successes.
    for n, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        title = str(vals[ti] or "").strip()
        author = str(vals[ai] or "").strip()
        isbn = clean_isbn(vals[ii]) if ii is not None else ""

        old = v4.get(n, {})
        variants = title_variants(title)

        safe = re.sub(
            r'[<>:"/\\|?*]',
            "_",
            f"{n:03d} - {title}"
        )[:180]
        path = OUTPUT_DIR / f"{safe}.jpg"

        if old.get("Status") == "Encontrada" and path.exists():
            results.append({
                "Nº": n,
                "Título original": title,
                "Variações de título": " | ".join(variants),
                "Autor": author,
                "ISBN": isbn,
                "Status": "Encontrada",
                "Arquivo": path.name,
                "Fonte": old.get("Fonte", ""),
                "URL capa": old.get("URL capa", ""),
                "Score": old.get("Score", ""),
                "Observação": "Preservada da v4",
            })
            continue

        print(f"[{n:03d}/{total}] {title} — {author}")
        if len(variants) > 1:
            print("        Busca também:", " | ".join(variants[:4]))

        status = "Não encontrada"
        source = ""
        url = ""
        score = ""
        observation = ""

        # 1. Retry the structured sources quickly.
        gurl, gsrc = google_cover(title, author, isbn)
        if gurl:
            ok, err = download_image(gurl, path)
            if ok:
                status = "Encontrada"
                source = gsrc
                url = gurl
                score = 100
                observation = "Fonte estruturada"
        if status != "Encontrada":
            ourl, osrc = openlibrary_cover(title, author, isbn)
            if ourl:
                ok, err = download_image(ourl, path)
                if ok:
                    status = "Encontrada"
                    source = osrc
                    url = ourl
                    score = 95
                    observation = "Fonte estruturada"

        # 2. General web search — the new layer.
        if status != "Encontrada":
            result = web_cover(title, author, isbn)
            if result:
                img_url, page_url, q, img_src = result
                ok, err = download_image(img_url, path)
                if ok:
                    status = "Encontrada"
                    source = "Busca web / " + domain(page_url)
                    url = img_url
                    score = round(q, 1)
                    observation = f"Página: {page_url} | imagem: {img_src}"
                else:
                    observation = f"Imagem encontrada, mas download falhou: {err}"

        if status == "Encontrada":
            print(f"        -> ENCONTRADA [{source}]")
        else:
            print("        -> NÃO ENCONTRADA")

        results.append({
            "Nº": n,
            "Título original": title,
            "Variações de título": " | ".join(variants),
            "Autor": author,
            "ISBN": isbn,
            "Status": status,
            "Arquivo": path.name if status == "Encontrada" else "",
            "Fonte": source,
            "URL capa": url,
            "Score": score,
            "Observação": observation,
        })

        # Public search engines are rate-sensitive.
        time.sleep(1.0)

    # Save report.
    rb = Workbook()
    rw = rb.active
    rw.title = "Resultado"

    cols = list(results[0].keys())
    rw.append(cols)
    for row in results:
        rw.append([row[c] for c in cols])

    sw = rb.create_sheet("Resumo")
    found = sum(x["Status"] == "Encontrada" for x in results)
    sw.append(["Total", total])
    sw.append(["Encontradas", found])
    sw.append(["Faltantes", total - found])
    sw.append([])
    sw.append(["Estratégia v5"])
    sw.append(["Google Books + Open Library + busca geral na web"])
    sw.append(["A v5 preserva as capas encontradas na v4 e tenta somente as faltantes."])

    rb.save(REPORT_FILE)

    with zipfile.ZipFile(ZIP_FILE, "w", zipfile.ZIP_DEFLATED) as z:
        for img in sorted(OUTPUT_DIR.glob("*.jpg")):
            z.write(img, img.name)

    print()
    print("=" * 72)
    print(f"CONCLUÍDO: {found}/{total}")
    print(f"Faltantes: {total - found}")
    print(f"ZIP:       {ZIP_FILE.resolve()}")
    print(f"RELATÓRIO: {REPORT_FILE.resolve()}")
    print("=" * 72)
    input("Pressione Enter para sair...")

if __name__ == "__main__":
    main()
