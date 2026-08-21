#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
COLETOR DE CAPAS v4

Mudança principal:
O título da planilha pode conter informações extras de série, volume,
número do livro etc. A v4 cria várias versões do título antes da busca.

Exemplo:
    "Os cinco porquinhos (Hercule Poirot, #25)"
vira:
    "Os cinco porquinhos"

Também tenta:
- texto antes de " - (..."
- texto antes de " (..."
- texto antes de " [..."
- remover "#25", "vol. 25", "volume 25", "book 25", etc.
- título completo e título simplificado
"""

import re
import time
import zipfile
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook, Workbook

INPUT = Path("Lista_de_Livros_com_genero.xlsx")
OUTPUT_DIR = Path("capas_livros_v4")
ZIP_FILE = Path("capas_livros_v4.zip")
REPORT_FILE = Path("resultado_capas_v4.xlsx")

GB_URL = "https://www.googleapis.com/books/v1/volumes"
OL_SEARCH = "https://openlibrary.org/search.json"
OL_COVER_ID = "https://covers.openlibrary.org/b/id/{cover_id}-L.jpg"
OL_COVER_ISBN = "https://covers.openlibrary.org/b/isbn/{isbn}-L.jpg"

session = requests.Session()
session.headers.update({
    "User-Agent": "BookCoverCollector/4.0 personal-library"
})

def clean_isbn(v):
    if v is None:
        return ""
    return re.sub(r"[^0-9Xx]", "", str(v).strip()).upper()

def isbn_variants(isbn):
    isbn = clean_isbn(isbn)
    out = []
    if isbn:
        out.append(isbn)
    if len(isbn) == 13 and isbn.startswith("978") and isbn[:-1].isdigit():
        core = isbn[3:-1]
        s = sum(int(ch) * (1 if i % 2 == 0 else 3)
                for i, ch in enumerate(core))
        out.append(core + str((10 - s % 10) % 10))
    elif len(isbn) == 10 and isbn[:9].isdigit():
        core = "978" + isbn[:9]
        s = sum(int(ch) * (1 if i % 2 == 0 else 3)
                for i, ch in enumerate(core))
        out.append(core + str((10 - s % 10) % 10))
    return list(dict.fromkeys(out))

def normalize_text(s):
    s = str(s or "").strip().lower()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()

def title_variants(title):
    """
    Retorna várias versões plausíveis do título, da mais conservadora
    para a mais agressiva.
    """
    original = str(title or "").strip()
    variants = []

    def add(x):
        x = re.sub(r"\s+", " ", str(x or "")).strip(" -–—:;,.")
        if len(x) >= 2 and x not in variants:
            variants.append(x)

    add(original)

    # Remove conteúdo entre parênteses quando ele parece ser série/volume.
    # Ex.: Os cinco porquinhos (Hercule Poirot, #25)
    s = re.sub(
        r"\s*[\(\[][^)\]]*(?:#\s*\d+|n[ºo°]?\s*\d+|volume|vol\.?|book|livro|s[ée]rie|series)[^)\]]*[\)\]]",
        "",
        original,
        flags=re.I
    )
    add(s)

    # Frequentíssimo em exports:
    # "Título - (Série, #5)"
    for sep in (" - (", " – (", " — (", " - [", " – [", " — ["):
        if sep in original:
            add(original.split(sep, 1)[0])

    # Se começa um parêntese/bracket, tudo antes normalmente é o título.
    m = re.search(r"\s*[\(\[]", original)
    if m:
        add(original[:m.start()])

    # Remover sufixos explícitos de volume/edição.
    s2 = re.sub(
        r"\s*[-–—:;,]?\s*[\(\[]?\s*"
        r"(?:livro|book|volume|vol\.?|n[ºo°]?|#)\s*\d+"
        r"[^\)\]]*[\)\]]?\s*$",
        "",
        original,
        flags=re.I
    )
    add(s2)

    # Remove "#25" ou "(#25)" no final.
    s3 = re.sub(r"\s*[\(\[]?\s*#\s*\d+\s*[\)\]]?\s*$", "", original)
    add(s3)

    # Remove "Hercule Poirot, #25" etc. se estiver entre parênteses.
    s4 = re.sub(r"\s*[\(\[][^)\]]*#\s*\d+[^)\]]*[\)\]]", "", original)
    add(s4)

    # Por fim, uma versão somente com a primeira parte antes de hífen
    # quando a segunda parte parece metadado.
    parts = re.split(r"\s+[-–—]\s+", original, maxsplit=1)
    if len(parts) == 2:
        rhs = normalize_text(parts[1])
        metadata_words = (
            "series", "série", "book", "livro", "volume", "vol",
            "crônicas", "cronicas", "trilogy", "trilogia", "saga"
        )
        if "#" in parts[1] or any(w in rhs for w in metadata_words):
            add(parts[0])

    return variants

def title_score(wanted, got):
    a, b = normalize_text(wanted), normalize_text(got)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return .94
    wa, wb = set(a.split()), set(b.split())
    return len(wa & wb) / max(1, len(wa | wb))

def author_score(wanted, got):
    if not wanted or not got:
        return 0.0

    def last_names(s):
        out = []
        for part in re.split(r"[,;&/]| and | e ", str(s), flags=re.I):
            x = normalize_text(part).split()
            if x:
                out.append(x[-1])
        return out

    a = last_names(wanted)
    b = []
    for person in got:
        x = normalize_text(person).split()
        if x:
            b.append(x[-1])

    if not a or not b:
        return 0.0
    return max(1.0 if x == y else 0.0 for x in a for y in b)

def google_search(query):
    try:
        r = session.get(
            GB_URL,
            params={"q": query, "maxResults": 40},
            timeout=25
        )
        if r.status_code != 200:
            return [], f"HTTP {r.status_code}: {r.text[:200]}"
        return r.json().get("items", []), ""
    except Exception as e:
        return [], f"{type(e).__name__}: {e}"

def best_google(items, wanted_titles, author, isbn):
    isbn_set = set(isbn_variants(isbn))
    best = None
    best_score = -1
    best_title = ""

    for item in items:
        vi = item.get("volumeInfo", {})
        if not (vi.get("imageLinks") or {}):
            continue

        ids = {
            clean_isbn(x.get("identifier"))
            for x in vi.get("industryIdentifiers", [])
            if x.get("identifier")
        }

        exact_isbn = bool(isbn_set & ids)
        ts = max(title_score(t, vi.get("title", "")) for t in wanted_titles)
        aus = author_score(author, vi.get("authors", []))

        # ISBN is extremely strong. Otherwise title+author decide.
        score = (250 if exact_isbn else 0) + 65 * ts + 45 * aus

        if score > best_score:
            best = item
            best_score = score
            best_title = vi.get("title", "")

    return best, best_score, best_title

def google_cover(title, author, isbn):
    variants = title_variants(title)

    # ISBN first.
    if isbn:
        for v in isbn_variants(isbn):
            items, err = google_search(f"isbn:{v}")
            if not err:
                item, score, found_title = best_google(
                    items, variants, author, isbn
                )
                if item:
                    return google_image(item), item, "Google Books / ISBN", score
            time.sleep(.15)

    # Then each title variation, first with author and then without.
    queries = []
    for t in variants:
        queries.append(f'intitle:"{t}" inauthor:"{author}"')
        queries.append(f'"{t}" "{author}"')
        queries.append(f'intitle:"{t}"')

    seen = set()
    for q in queries:
        if q in seen:
            continue
        seen.add(q)

        items, err = google_search(q)
        if err:
            continue

        item, score, found_title = best_google(
            items, variants, author, isbn
        )

        # 50 allows title-only matches when the author isn't indexed exactly.
        if item and score >= 50:
            return google_image(item), item, "Google Books / variação do título", score

        time.sleep(.15)

    return None, None, "Google Books: nenhum candidato", 0

def google_image(item):
    links = item.get("volumeInfo", {}).get("imageLinks") or {}
    for key in ("extraLarge", "large", "medium", "small", "thumbnail"):
        if links.get(key):
            return links[key].replace("http://", "https://")
    return None

def ol_search(params):
    try:
        r = session.get(OL_SEARCH, params=params, timeout=25)
        if r.status_code != 200:
            return [], f"HTTP {r.status_code}: {r.text[:200]}"
        return r.json().get("docs", []), ""
    except Exception as e:
        return [], f"{type(e).__name__}: {e}"

def best_ol(docs, wanted_titles, author, isbn):
    isbn_set = set(isbn_variants(isbn))
    best = None
    best_score = -1

    for doc in docs:
        cover_id = doc.get("cover_i")
        if not cover_id:
            continue

        ids = set()
        for key in ("isbn", "isbn13", "isbn10"):
            vals = doc.get(key) or []
            if isinstance(vals, str):
                vals = [vals]
            ids.update(clean_isbn(x) for x in vals)

        exact = bool(isbn_set & ids)
        ts = max(title_score(t, doc.get("title", "")) for t in wanted_titles)
        aus = author_score(author, doc.get("author_name", []))

        score = (250 if exact else 0) + 65 * ts + 45 * aus

        if score > best_score:
            best = doc
            best_score = score

    return best, best_score

def openlibrary_cover(title, author, isbn):
    variants = title_variants(title)

    # ISBN searches.
    for v in isbn_variants(isbn):
        docs, err = ol_search({"isbn": v, "limit": 40})
        if not err:
            doc, score = best_ol(docs, variants, author, isbn)
            if doc and doc.get("cover_i"):
                return (
                    OL_COVER_ID.format(cover_id=doc["cover_i"]),
                    doc,
                    "Open Library / ISBN",
                    score
                )
        time.sleep(.2)

    # Search each cleaned title.
    for t in variants:
        queries = [
            {"title": t, "author": author, "limit": 40},
            {"q": f'"{t}" "{author}"', "limit": 40},
            {"title": t, "limit": 40},
        ]

        for params in queries:
            docs, err = ol_search(params)
            if err:
                continue

            doc, score = best_ol(docs, variants, author, isbn)
            if doc and doc.get("cover_i") and score >= 50:
                return (
                    OL_COVER_ID.format(cover_id=doc["cover_i"]),
                    doc,
                    "Open Library / variação do título",
                    score
                )
            time.sleep(.15)

    # Direct ISBN is only the final fallback.
    for v in isbn_variants(isbn):
        url = OL_COVER_ISBN.format(isbn=v) + "?default=false"
        try:
            r = session.get(url, timeout=20)
            if r.status_code == 200 and len(r.content) > 5000:
                return url, {}, "Open Library / ISBN direto", 999
        except Exception:
            pass

    return None, None, "Open Library: nenhum candidato", 0

def download(url, path):
    try:
        r = session.get(url, timeout=35)
        if r.status_code != 200:
            return False, f"HTTP {r.status_code}"

        data = r.content
        ct = r.headers.get("content-type", "").lower()

        if len(data) < 5000:
            return False, f"arquivo muito pequeno ({len(data)} bytes)"

        if "image" not in ct and not data.startswith(
            (b"\xff\xd8", b"\x89PNG", b"RIFF")
        ):
            return False, f"não parece imagem ({ct})"

        path.write_bytes(data)
        return True, ""
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"

def find_columns(ws):
    headers = [str(c.value or "").strip() for c in ws[1]]

    def find(names):
        for name in names:
            for i, h in enumerate(headers):
                if normalize_text(h) == normalize_text(name):
                    return i
        return None

    ti = find(["Título", "Titulo", "Title"])
    ai = find(["Autor", "Author"])
    ii = find(["ISBN", "ISBN13", "ISBN-13", "ISBN10", "ISBN-10"])

    if ti is None or ai is None:
        raise RuntimeError(f"Colunas encontradas: {headers}")

    return ti, ai, ii

def main():
    print("=" * 70)
    print("COLETOR DE CAPAS v4 — VARIAÇÕES DE TÍTULO")
    print("=" * 70)
    print()

    if not INPUT.exists():
        print(f"ERRO: não encontrei {INPUT.resolve()}")
        input("Pressione Enter para sair...")
        return

    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = load_workbook(INPUT, data_only=True)
    ws = wb.active
    ti, ai, ii = find_columns(ws)

    total = ws.max_row - 1
    results = []

    for n, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        title = str(vals[ti] or "").strip()
        author = str(vals[ai] or "").strip()
        isbn = clean_isbn(vals[ii]) if ii is not None else ""

        variants = title_variants(title)

        safe = re.sub(
            r'[<>:"/\\|?*]',
            "_",
            f"{n:03d} - {title}"
        )[:180]

        path = OUTPUT_DIR / f"{safe}.jpg"

        print(f"[{n:03d}/{total}] {title} — {author}")
        if len(variants) > 1:
            print(f"        Títulos testados: {' | '.join(variants[:4])}")

        status = "Não encontrada"
        source = ""
        url = ""
        error = ""
        found_title = ""
        found_author = ""
        score = 0

        gurl, gitem, gsource, gscore = google_cover(
            title, author, isbn
        )

        if gurl:
            ok, derr = download(gurl, path)
            if ok:
                status = "Encontrada"
                source = gsource
                url = gurl
                score = gscore
                vi = gitem.get("volumeInfo", {}) if gitem else {}
                found_title = vi.get("title", "")
                found_author = ", ".join(vi.get("authors", []))
            else:
                error = f"{gsource}: {derr}"

        if status != "Encontrada":
            ourl, odoc, osource, oscore = openlibrary_cover(
                title, author, isbn
            )

            if ourl:
                ok, derr = download(ourl, path)
                if ok:
                    status = "Encontrada"
                    source = osource
                    url = ourl
                    score = oscore

                    if odoc:
                        found_title = odoc.get("title", "")
                        found_author = ", ".join(
                            odoc.get("author_name", [])
                        )
                else:
                    error += (
                        (" | " if error else "")
                        + f"{osource}: {derr}"
                    )
            else:
                error += (
                    (" | " if error else "")
                    + osource
                )

        if status == "Encontrada":
            print(
                f"        -> ENCONTRADA [{source}] "
                f"(score {score:.0f})"
            )
        else:
            print(f"        -> NÃO ENCONTRADA")
            print(f"        -> {error}")

        results.append({
            "Nº": n,
            "Título original": title,
            "Variações de título usadas": " | ".join(variants),
            "Autor": author,
            "ISBN": isbn,
            "Status": status,
            "Arquivo": path.name if status == "Encontrada" else "",
            "Fonte": source,
            "Score": round(score, 2),
            "Título encontrado": found_title,
            "Autor encontrado": found_author,
            "URL capa": url,
            "Erro/observação": error,
        })

        time.sleep(.4)

    rb = Workbook()
    rw = rb.active
    rw.title = "Resultado"

    cols = list(results[0].keys())
    rw.append(cols)

    for row in results:
        rw.append([row[c] for c in cols])

    sw = rb.create_sheet("Resumo")
    found = sum(x["Status"] == "Encontrada" for x in results)
    sw.append(["Total", total])
    sw.append(["Capas encontradas", found])
    sw.append(["Capas faltantes", total - found])
    sw.append([])
    sw.append(["Exemplo de limpeza"])
    sw.append([
        "Os cinco porquinhos (Hercule Poirot, #25)",
        "Os cinco porquinhos"
    ])

    rb.save(REPORT_FILE)

    with zipfile.ZipFile(
        ZIP_FILE,
        "w",
        zipfile.ZIP_DEFLATED
    ) as z:
        for img in sorted(OUTPUT_DIR.glob("*.jpg")):
            z.write(img, img.name)

    print()
    print("=" * 70)
    print(f"CONCLUÍDO: {found}/{total}")
    print(f"ZIP:       {ZIP_FILE.resolve()}")
    print(f"RELATÓRIO: {REPORT_FILE.resolve()}")
    print("=" * 70)
    input("Pressione Enter para sair...")

if __name__ == "__main__":
    main()
